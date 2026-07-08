import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

EXPORT_DIR   = r"C:\Users\Admin\Desktop\Отчеты"

# Цвета
C_HDR_BLUE   = "1B3A6B"
C_HDR_DARK   = "0D3B6E"
C_BORDER     = "B0B8C8"
C_ODD        = "F5F7FA"
C_EVEN       = "FFFFFF"
C_LINK       = "1565C0"
C_DONE       = "25D366"
C_DONE_TXT   = "FFFFFF"
C_INWORK     = "90CAF9"
C_INWORK_TXT = "0D2B5E"
C_TALKS      = "FFF176"
C_TALKS_TXT  = "4A3A00"
C_REVISION   = "FFCC80"
C_REVISION_TXT = "5E3500"
C_REJECTED   = "EF9A9A"
C_REJECTED_TXT = "5B0000"
C_TOTAL_BG   = "1B3A6B"
C_TOTAL_TXT  = "C9A227"

STATUS_COLORS = {
    "Выполнено":       (C_DONE,     C_DONE_TXT),
    "В работе":        (C_INWORK,   C_INWORK_TXT),
    "Переговоры":      (C_TALKS,    C_TALKS_TXT),
    "На доработке":    (C_REVISION, C_REVISION_TXT),
    "Отказ заказчика": (C_REJECTED, C_REJECTED_TXT),
    "Отклонили":       (C_REJECTED, C_REJECTED_TXT),
}


def _report_path_for_date(date_str: str) -> str:
    """
    Возвращает путь к файлу отчёта по дате создания записи.
    date_str — дата в любом формате, из которой извлекаем ДД.ММ.ГГГГ.
    Изменения в задании от 01.07 всегда идут в Отчет_01.07.2026.xlsx,
    независимо от того, какое сегодня число.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)
    # Пробуем распарсить дату из строки вида "01.07.2026 20:48" или "2026-07-01 20:48"
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(date_str[:16], fmt[:len(fmt)])
            return os.path.join(EXPORT_DIR, dt.strftime("Отчет_%d.%m.%Y.xlsx"))
        except ValueError:
            continue
    # Если не удалось распарсить — используем сегодня
    return os.path.join(EXPORT_DIR,
                        datetime.now().strftime("Отчет_%d.%m.%Y.xlsx"))


def _border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _hdr(cell, text, bg=C_HDR_DARK, size=10):
    cell.value = text
    cell.font  = Font(bold=True, color="FFFFFF", size=size)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = _center()
    cell.border = _border("FFFFFF")


def _cell(cell, value, bg=C_EVEN, fg="000000", bold=False, size=9,
          wrap=True, link=None):
    cell.value = value
    cell.font  = Font(color=fg, bold=bold, size=size)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = _center()
    cell.border = _border()
    if link:
        cell.hyperlink = link
        cell.font = Font(color=C_LINK, underline="single", size=size)


def _status_colors(status_raw: str):
    for kw, (bg, fg) in STATUS_COLORS.items():
        if kw in status_raw:
            return bg, fg
    return C_EVEN, "000000"


# ─── Колонки отчёта ──────────────────────────────────────────────────────────
REPORT_COLS = [
    ("Дата / Время",      16),
    ("Платформа",         13),
    ("Канал / Группа",    22),
    ("Категория",         24),
    ("Схожесть %",        11),
    ("Превью задания",    48),
    ("Ссылка",            38),
    ("Статус",            22),
    ("Дата изменения",    18),   # ← новый столбец
    ("Сумма (₽)",         13),
    ("Способ оплаты",     22),
    ("Комментарий",       35),
]
N_COLS = len(REPORT_COLS)


def _build_fresh_sheet(ws):
    """Создаёт шапку таблицы на листе"""
    ws.sheet_view.showGridLines = False
    for c, (header, width) in enumerate(REPORT_COLS, 1):
        _hdr(ws.cell(row=1, column=c), header)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, item: dict):
    """Записывает/перезаписывает одну строку данных"""
    status_raw = item.get("status", "")
    bg, fg = _status_colors(status_raw)
    row_bg = bg if bg != C_EVEN else (C_ODD if row_idx % 2 == 0 else C_EVEN)

    def c(col, val, **kw):
        _cell(ws.cell(row_idx, col), val, row_bg, fg, **kw)

    c(1,  item.get("date", ""))
    # Платформа — жирная, чуть крупнее
    _cell(ws.cell(row_idx, 2), item.get("platform", ""),
          row_bg, fg, bold=True, size=10)
    # Канал/группа — жирный
    _cell(ws.cell(row_idx, 3), item.get("channel", ""),
          row_bg, fg, bold=True)
    c(4,  item.get("category", ""))
    c(5,  item.get("match_percent", 0))
    c(6,  item.get("text_preview", ""))
    link = item.get("link", "")
    _cell(ws.cell(row_idx, 7), link, row_bg, fg, link=link if link else None)
    c(8,  status_raw)
    c(9,  item.get("changed_at", ""))    # Дата изменения

    amount_str = str(item.get("amount", "")).replace(" ", "").replace("₽", "")
    try:
        amount = float(amount_str) if amount_str else 0.0
    except ValueError:
        amount = 0.0
    amt_cell = ws.cell(row_idx, 10)
    _cell(amt_cell, amount if amount else "", row_bg, fg)
    if amount:
        amt_cell.number_format = '#,##0 ₽'

    c(11, item.get("payment_value", ""))
    c(12, item.get("comment", ""))
    ws.row_dimensions[row_idx].height = 42


def _find_row_by_key(ws, key: str) -> int:
    """Ищет строку по ключу в скрытом примечании ячейки A / по совпадению данных"""
    # Ключ хранится в ячейке A (комментарий не поддерживается openpyxl просто)
    # Храним key как текст в столбце M (скрытый вспомогательный)
    for row in range(2, ws.max_row + 1):
        cell_key = ws.cell(row, N_COLS + 1).value  # столбец M
        if cell_key == key:
            return row
    return -1


def export_results_report(results_data: dict) -> str:
    """
    Выгружает отчёт результатов.
    Каждая запись идёт в файл по дате своего первого создания (report_date / date).
    Меняешь статус задания от 01.07 в день 02.07 → изменения в Отчет_01.07.2026.xlsx.
    Возвращает путь к файлу последней изменённой записи.
    """
    # Группируем записи по дате файла (дата первого создания записи)
    by_file: dict[str, dict] = {}
    for key, item in results_data.items():
        # report_date — дата создания записи, по ней находим нужный файл
        record_date = item.get("report_date") or item.get("date", "")
        filepath    = _report_path_for_date(record_date)
        if filepath not in by_file:
            by_file[filepath] = {}
        by_file[filepath][key] = item

    last_path = os.path.join(EXPORT_DIR,
                             datetime.now().strftime("Отчет_%d.%m.%Y.xlsx"))

    for filepath, file_items in by_file.items():
        last_path = filepath

        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчёт по результатам"
            _build_fresh_sheet(ws)
            ws.column_dimensions[get_column_letter(N_COLS + 1)].hidden = True
            ws.column_dimensions[get_column_letter(N_COLS + 1)].width  = 1

        total_amount = 0.0

        for key, item in file_items.items():
            existing_row = _find_row_by_key(ws, key)
            if existing_row > 0:
                _write_row(ws, existing_row, item)
                ws.cell(existing_row, N_COLS + 1).value = key
            else:
                new_row = ws.max_row + 1
                _write_row(ws, new_row, item)
                ws.cell(new_row, N_COLS + 1).value = key

            amount_str = str(item.get("amount", "")).replace(" ", "").replace("₽", "")
            try:
                total_amount += float(amount_str) if amount_str else 0.0
            except ValueError:
                pass

        # Итоговая строка — удаляем старую и добавляем новую
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 8).value == "ИТОГО:":
                ws.delete_rows(row)
                break

        tr = ws.max_row + 1
        for col in range(1, N_COLS + 2):
            ws.cell(tr, col).fill   = PatternFill("solid", fgColor=C_TOTAL_BG)
            ws.cell(tr, col).border = _border("FFFFFF")
        ws.cell(tr, 8).value     = "ИТОГО:"
        ws.cell(tr, 8).font      = Font(bold=True, color=C_TOTAL_TXT, size=11)
        ws.cell(tr, 8).alignment = _center()
        tc = ws.cell(tr, 10)
        tc.value         = total_amount
        tc.number_format = '#,##0 ₽'
        tc.font          = Font(bold=True, color=C_TOTAL_TXT, size=11)
        tc.alignment     = _center()
        ws.row_dimensions[tr].height = 28

        wb.save(filepath)

    return os.path.abspath(last_path)




# ─── Выгрузка найденных заявок (дневной архив) ───────────────────────────────

def export_to_excel(results: list, filepath: str = None) -> str:
    """Ежедневный архив найденных заявок — отдельный файл"""
    if not filepath:
        date_str = datetime.now().strftime("%d.%m.%Y_%H-%M")
        os.makedirs(EXPORT_DIR, exist_ok=True)
        filepath = os.path.join(EXPORT_DIR, f"Заявки_{date_str}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Найденные заявки"
    ws.sheet_view.showGridLines = False

    cols = [
        ("Дата / Время",   16), ("Платформа",  13), ("Канал / Группа", 22),
        ("Категория",      24), ("Схожесть %", 11), ("Текст заявки",   55),
        ("Ссылка",         38), ("Статус",     14), ("Комментарий",    30),
    ]
    for c, (h, w) in enumerate(cols, 1):
        _hdr(ws.cell(row=1, column=c), h, bg=C_HDR_BLUE)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for r, item in enumerate(results, 2):
        bg = C_ODD if r % 2 == 0 else C_EVEN
        _cell(ws.cell(r, 1), datetime.now().strftime("%d.%m.%Y %H:%M"), bg)
        _cell(ws.cell(r, 2), item.get("platform", ""), bg, bold=True, size=10)
        _cell(ws.cell(r, 3), item.get("channel", ""),  bg, bold=True)
        _cell(ws.cell(r, 4), item.get("category", ""), bg)
        _cell(ws.cell(r, 5), item.get("match_percent", 0), bg)
        _cell(ws.cell(r, 6), item.get("text", "")[:500], bg)
        link = item.get("link", "")
        _cell(ws.cell(r, 7), link, bg, link=link if link else None)
        _cell(ws.cell(r, 8), item.get("status", "новый"), bg)
        _cell(ws.cell(r, 9), item.get("reply", "")[:200], bg)
        ws.row_dimensions[r].height = 40

    wb.save(filepath)
    return os.path.abspath(filepath)